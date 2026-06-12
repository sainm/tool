"""
JDK API 导出到 Excel，支持多版本对比。

用法:
    python export_api_to_excel.py --jdk /path/to/jdk-17 --out jdk17.xlsx
    python export_api_to_excel.py --jdk /path/to/jdk-21 --diff-with jdk17.xlsx --out diff_17vs21.xlsx

依赖:
    pip install openpyxl
"""

import argparse
import subprocess
import sys
import os
import re
import tempfile
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# ============================================================
# 样式定义
# ============================================================
HEADER_FILL = PatternFill(start_color="4D7A97", end_color="4D7A97", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ADD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")      # 绿色
REMOVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # 红色
CHANGE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")    # 蓝色
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============================================================
# API 数据模型
# ============================================================
class Member:
    __slots__ = ('kind', 'name', 'signature', 'modifiers')
    def __init__(self, kind, name, signature, modifiers=""):
        self.kind = kind           # "constructor", "method", "field"
        self.name = name
        self.signature = signature # 完整签名，如 "void foo(int)"
        self.modifiers = modifiers # "public static final" 等

    def key(self):
        return (self.kind, self.name, self.signature)

class TypeInfo:
    __slots__ = ('module', 'pkg', 'kind', 'name', 'members')
    def __init__(self, module, pkg, kind, name):
        self.module = module
        self.pkg = pkg
        self.kind = kind      # "class", "interface", "enum", "record", "annotation"
        self.name = name
        self.members = []     # list of Member

    def full_name(self):
        return f"{self.pkg}.{self.name}"


def find_java_bin(jdk_home):
    """查找 java / javap 可执行文件"""
    java_exe = Path(jdk_home) / "bin" / "java.exe"
    javap_exe = Path(jdk_home) / "bin" / "javap.exe"
    if not java_exe.exists():
        java_exe = Path(jdk_home) / "bin" / "java"
        javap_exe = Path(jdk_home) / "bin" / "javap"
    return str(java_exe), str(javap_exe)


def list_modules(java_exe):
    """列出 JDK 所有模块名"""
    result = subprocess.run(
        [java_exe, "--list-modules"],
        capture_output=True, text=True, timeout=60
    )
    modules = []
    for line in result.stdout.strip().splitlines():
        mod = line.strip().split("@")[0]
        if mod:  # 跳过空行、jdk.jdeps 内部模块等
            modules.append(mod)
    return modules


def list_module_classes(java_exe, module):
    """列出某个模块中所有公开的 .class 文件（通过 jrt: 文件系统）"""
    code = f'''
import java.nio.file.*;
import java.util.stream.*;
import java.io.IOException;
public class Ls {{
    public static void main(String[] args) throws Exception {{
        var fs = FileSystems.getFileSystem(java.net.URI.create("jrt:/"));
        var classes = Files.walk(fs.getPath("modules/{module}"))
            .filter(p -> p.toString().endsWith(".class"))
            .map(p -> p.toString()
                .replaceFirst("^modules/{module}/", "")
                .replace(".class", ""))
            .filter(n -> !n.contains("$"))
            .sorted()
            .collect(java.util.stream.Collectors.toList());
        for (var c : classes) System.out.println(c);
    }}
}}
    '''
    with tempfile.NamedTemporaryFile(mode='w', suffix='.java', delete=False, encoding='utf-8') as f:
        java_file = f.name
        f.write(code)

    # 编译
    subprocess.run(
        [java_exe.replace("java", "javac") if "javac" in java_exe else java_exe.replace("java.exe", "javac.exe"), java_file],
        capture_output=True, text=True, timeout=30
    )

    # 运行
    result = subprocess.run(
        [java_exe, "-cp", os.path.dirname(java_file), "Ls", module],
        capture_output=True, text=True, timeout=120
    )
    os.unlink(java_file)
    for cls in ("Ls.class", "Ls.java"):
        tmp_cls = Path(os.path.dirname(java_file)) / cls
        if tmp_cls.exists():
            tmp_cls.unlink()

    if result.returncode != 0:
        print(f"  [警告] 列出模块 {module} 失败: {result.stderr.strip()}")
        return []
    return [l.strip() for l in result.stdout.strip().splitlines() if l.strip()]


def parse_javap_output(text, module, pkg, cls_name):
    """解析 javap -public 输出"""
    # 判断类型
    kind = "class"
    if text.startswith("public interface "):
        kind = "interface"
    elif text.startswith("public enum "):
        kind = "enum"
    elif text.startswith("public record "):
        kind = "record"
    elif text.startswith("public @interface "):
        kind = "annotation"

    ti = TypeInfo(module=module, pkg=pkg, kind=kind, name=cls_name)

    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # 跳过文件头（Compiled from...）
        if line.startswith("Compiled from"):
            i += 1
            continue

        # 构造函数
        m = re.match(r'^\s*(public\s+)(\w[\w<>,\s\[\]\.]*)\s*\((.*)\)\s*(?:throws\s+.*)?\s*;\s*$', line)
        if m and '(' in line and ')' in line:
            name_part = m.group(2).strip()
            params = m.group(3).strip()
            # 判断是构造函数还是普通方法：名字不含返回类型
            # 如果是构造函数，行首就是 "public ClassName("
            ctor_m = re.match(r'public\s+(\w[\w<>,\s\[\]\.]*)\s*\((.*)\)\s*', line)
            method_m = re.match(r'public\s+(static\s+)?(\w[\w<>,\s\[\]\.]*)\s+(\w[\w<>]*)\((.*)\)\s*', line)
            if method_m:
                mods = method_m.group(1) or ""
                ret_type = method_m.group(2).strip()
                m_name = method_m.group(3).strip()
                m_params = method_m.group(4).strip()
                ti.members.append(Member(
                    kind="method",
                    name=m_name,
                    signature=f"{mods}{ret_type} {m_name}({m_params})".strip(),
                    modifiers=f"public {mods}".strip()
                ))
            elif ctor_m and cls_name in line:
                ti.members.append(Member(
                    kind="constructor",
                    name=cls_name,
                    signature=f"{cls_name}({params})"
                ))
            i += 1
            continue

        # 字段
        f_m = re.match(r'^\s*public\s+(static\s+)?(final\s+)?(\w[\w<>,\s\[\]\.]*)\s+(\w+)\s*;?\s*$', line)
        if f_m and '{' not in line and '(' not in line:
            ti.members.append(Member(
                kind="field",
                name=f_m.group(4),
                signature=line.strip().rstrip(';').strip(),
                modifiers="public"
            ))
            i += 1
            continue

        i += 1

    return ti


def extract_jdk_api(jdk_home, progress_callback=None):
    """提取一个 JDK 的全部公开 API"""
    java_exe, javap_exe = find_java_bin(jdk_home)
    print(f"\n[扫描 JDK] {jdk_home}")
    print(f"  java: {java_exe}")

    modules = list_modules(java_exe)
    print(f"  共 {len(modules)} 个模块")

    all_types = []
    total = len(modules)

    for idx, mod in enumerate(modules):
        if progress_callback:
            progress_callback(idx, total, mod)

        classes = list_module_classes(java_exe, mod)
        if not classes:
            continue

        for cls_path in classes:
            # cls_path 如 "java/lang/String"
            *pkg_parts, cls_name = cls_path.split("/")
            pkg = ".".join(pkg_parts) if pkg_parts else ""

            full = f"{pkg}.{cls_name}" if pkg else cls_name if cls_name else ""

            try:
                result = subprocess.run(
                    [javap_exe, "-public", "-classpath", "", f"--module={mod}", full],
                    capture_output=True, text=True, timeout=30
                )
                if result.returncode == 0 and result.stdout.strip():
                    ti = parse_javap_output(result.stdout, mod, pkg, cls_name)
                    all_types.append(ti)
            except Exception:
                continue

    print(f"  共提取 {len(all_types)} 个类型")
    return all_types


# ============================================================
# Excel 写入
# ============================================================

def write_api_sheet(ws, types, title="API"):
    """把一个 JDK 的 API 写入一个 sheet"""
    # 表头
    headers = ["模块", "包", "类型名", "类型", "成员类型", "成员名", "成员签名"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    row = 2
    for ti in sorted(types, key=lambda x: (x.module, x.pkg, x.name)):
        if not ti.members:
            # 没有成员也列出来（纯标记类型或空接口）
            ws.cell(row=row, column=1, value=ti.module).border = THIN_BORDER
            ws.cell(row=row, column=2, value=ti.pkg).border = THIN_BORDER
            ws.cell(row=row, column=3, value=ti.name).border = THIN_BORDER
            ws.cell(row=row, column=4, value=ti.kind).border = THIN_BORDER
            row += 1
        else:
            for m in ti.members:
                ws.cell(row=row, column=1, value=ti.module).border = THIN_BORDER
                ws.cell(row=row, column=2, value=ti.pkg).border = THIN_BORDER
                ws.cell(row=row, column=3, value=ti.name).border = THIN_BORDER
                ws.cell(row=row, column=4, value=ti.kind).border = THIN_BORDER
                ws.cell(row=row, column=5, value=m.kind).border = THIN_BORDER
                ws.cell(row=row, column=6, value=m.name).border = THIN_BORDER
                ws.cell(row=row, column=7, value=m.signature).border = THIN_BORDER
                row += 1

    # 冻结首行，自动宽度
    ws.freeze_panes = 'A2'
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22

    return row - 2  # 返回行数


def build_api_index(types):
    """把 TypeInfo 列表构建成查找表: (模块, 包, 类型名, 成员key) -> 成员签名"""
    index = {}
    type_index = {}  # (模块, 包, 类型名) -> 类型信息
    for ti in types:
        tkey = (ti.module, ti.pkg, ti.name)
        type_index[tkey] = ti
        for m in ti.members:
            index[(ti.module, ti.pkg, ti.name, m.kind, m.name, m.signature)] = m
    return index, type_index


def write_diff_sheet(ws, base_types, new_types, base_label="旧", new_label="新"):
    """比较两个 JDK，写入差异 sheet"""
    headers = ["状态", "模块", "包", "类型名", "类型", "成员类型", "成员名",
               f"{base_label}签名", f"{new_label}签名"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    base_idx, base_types_idx = build_api_index(base_types)
    new_idx, new_types_idx = build_api_index(new_types)

    all_keys = set(base_idx.keys()) | set(new_idx.keys())
    # 也收集类型级别的 key
    all_type_keys = set(base_types_idx.keys()) | set(new_types_idx.keys())

    row = 2
    diff_count = 0

    for tkey in sorted(all_type_keys):
        bt = base_types_idx.get(tkey)
        nt = new_types_idx.get(tkey)
        if bt is None:
            # 新增类型
            for m in nt.members:
                _write_diff_row(ws, row, "新增", nt, m, None, None)
                row += 1
                diff_count += 1
        elif nt is None:
            # 删除类型
            for m in bt.members:
                _write_diff_row(ws, row, "删除", None, None, bt, m)
                row += 1
                diff_count += 1
        # 类型都存在就不在类型级标记，交给成员级

    # 成员级对比
    member_keys = defaultdict(list)
    for k in all_keys:
        member_keys[(k[0], k[1], k[2])].append(k)

    for tkey in sorted(member_keys.keys()):
        bt = base_types_idx.get(tkey)
        nt = new_types_idx.get(tkey)
        if bt is None or nt is None:
            continue  # 类型新增/删除在上面的类型级循环已处理

        bt_members = {m.key(): m for m in bt.members}
        nt_members = {m.key(): m for m in nt.members}

        all_m_keys = set(bt_members.keys()) | set(nt_members.keys())

        for mk in sorted(all_m_keys):
            bm = bt_members.get(mk)
            nm = nt_members.get(mk)

            if bm and nm and bm.signature == nm.signature:
                continue  # 相同，不输出（可选：加 --show-unchanged 输出）

            if bm and not nm:
                _write_diff_row(ws, row, "删除", bt, bm, bt, bm)
                row += 1
                diff_count += 1
            elif not bm and nm:
                _write_diff_row(ws, row, "新增", nt, nm, None, None)
                row += 1
                diff_count += 1
            elif bm and nm:
                _write_diff_row(ws, row, "修改", nt, nm, bt, bm)
                row += 1
                diff_count += 1

    # 冻结首行
    ws.freeze_panes = 'A2'
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26

    return diff_count


def _write_diff_row(ws, row, status, new_ti, new_m, base_ti, base_m):
    """写入一行差异"""
    if status == "新增":
        fill = ADD_FILL
    elif status == "删除":
        fill = REMOVE_FILL
    else:
        fill = CHANGE_FILL

    ti = new_ti or base_ti
    data = [
        status,
        ti.module,
        ti.pkg,
        ti.name,
        ti.kind,
        (new_m or base_m).kind,
        (new_m or base_m).name,
        base_m.signature if base_m else "",
        new_m.signature if new_m else "",
    ]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER


def write_summary_sheet(ws, base_types, new_types, base_label, new_label):
    """写入汇总统计"""
    headers = ["统计项", base_label, new_label, "差异"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    base_idx, _ = build_api_index(base_types)
    new_idx, _ = build_api_index(new_types)

    base_count = len(base_types)
    new_count = len(new_types)
    base_member_count = len(base_idx)
    new_member_count = len(new_idx)

    all_keys = set(base_idx.keys()) | set(new_idx.keys())
    added = sum(1 for k in all_keys if k not in base_idx)
    removed = sum(1 for k in all_keys if k not in new_idx)
    changed = 0
    for k in all_keys:
        if k in base_idx and k in new_idx:
            if base_idx[k].signature != new_idx[k].signature:
                changed += 1

    rows = [
        ("类型总数", base_count, new_count, new_count - base_count),
        ("成员总数 (方法/字段/构造器)", base_member_count, new_member_count, new_member_count - base_member_count),
        ("新增成员", "-", added, added),
        ("删除成员", removed, "-", -removed),
        ("签名修改", "-", "-", changed),
    ]
    for r, (label, b, n, d) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws.cell(row=r, column=2, value=b).border = THIN_BORDER
        ws.cell(row=r, column=3, value=n).border = THIN_BORDER
        ws.cell(row=r, column=4, value=d).border = THIN_BORDER

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 28


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="JDK API 导出到 Excel")
    parser.add_argument("--jdk", required=True, help="JDK 安装目录")
    parser.add_argument("--out", required=True, help="输出 Excel 文件路径")
    parser.add_argument("--diff-with", help="与另一个 JDK 的 API Excel 比较，写入 diff sheet")
    parser.add_argument("--label", help="当前 JDK 的标签（比较时使用）", default="新版本")
    parser.add_argument("--base-label", help="基准 JDK 的标签（比较时使用）", default="旧版本")
    args = parser.parse_args()

    jdk_home = args.jdk
    if not os.path.isdir(jdk_home):
        print(f"错误: JDK 目录不存在: {jdk_home}")
        sys.exit(1)

    types = extract_jdk_api(jdk_home)

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    ws_api = wb.create_sheet("API列表")
    count = write_api_sheet(ws_api, types)
    print(f"\n[API列表] 写入 {count} 行")

    if args.diff_with:
        # 从已有 xlsx 读取基准数据
        base_label = args.base_label
        new_label = args.label

        if not os.path.exists(args.diff_with):
            print(f"警告: 基准文件不存在: {args.diff_with}, 跳过对比")
        else:
            from openpyxl import load_workbook
            base_wb = load_workbook(args.diff_with, read_only=True)
            base_ws = base_wb["API列表"]

            base_types = []
            current_type = None
            current_members = []
            for row in base_ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                mod, pkg, cls_name, kind, m_kind, m_name, m_sig = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
                ti_key = (mod, pkg, cls_name)
                if current_type is None or current_type.full_name() != f"{pkg}.{cls_name}" or current_type.module != mod:
                    if current_type is not None:
                        current_type.members = current_members
                        base_types.append(current_type)
                    current_type = TypeInfo(module=mod, pkg=pkg, kind=kind or "class", name=cls_name)
                    current_members = []
                if m_kind:
                    current_members.append(Member(kind=m_kind, name=m_name, signature=m_sig))
            if current_type is not None:
                current_type.members = current_members
                base_types.append(current_type)
            base_wb.close()

            print(f"  读取基准 API: {len(base_types)} 个类型")

            ws_diff = wb.create_sheet(f"对比 ({base_label} vs {new_label})")
            diff_count = write_diff_sheet(ws_diff, base_types, types, base_label, new_label)
            print(f"[对比] 写入 {diff_count} 行差异")

            ws_sum = wb.create_sheet("汇总")
            write_summary_sheet(ws_sum, base_types, types, base_label, new_label)

    wb.save(args.out)
    print(f"\n完成! 输出文件: {args.out}")


if __name__ == "__main__":
    main()
